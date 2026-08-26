"""
Tool: UCP Format PDF
Group Excel rows by Prospect No (or any column), format each group as a
styled Excel sheet with a custom header, export to PDF, then merge.
"""

import os
import re
import threading
import subprocess
from datetime import datetime
from pathlib import Path

import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfMerger

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

DESKTOP  = os.path.join(os.path.expanduser("~"), "Desktop")
BASE_OUT = os.path.join(DESKTOP, "OUTPUT", "UCP_Format_PDF")

DEFAULT_GROUP_COL = "Prospect_No"
DEFAULT_WIDE_COL = "Address"
DEFAULT_HEADER = "IIFL | Prospect No - {prospect_no}"
DEFAULT_BATCH = 500
WIDE_WIDTH = 70

C = {
    "bg":     "#0a0a0f", "card":   "#16161f", "hover":  "#1e1e2e",
    "border": "#2a2a3d", "text":   "#e8e8f0", "muted":  "#8888aa",
    "faint":  "#44445a", "accent": "#ff9f0a", "green":  "#30d158",
    "red":    "#ff375f",
}
TINT = {"bg": "#2e1e00", "mid": "#4a3000", "bdr": "#6b4500"}


def get_output_dir():
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = os.path.join(BASE_OUT, ts)
    os.makedirs(path, exist_ok=True)
    return path


def find_column(columns, preferred_names):
    """Case-insensitive / partial match for a column name."""
    lower_map = {str(c).strip().lower(): c for c in columns}
    for name in preferred_names:
        key = name.lower()
        if key in lower_map:
            return lower_map[key]
    for name in preferred_names:
        key = name.lower()
        for col_l, col in lower_map.items():
            if key in col_l:
                return col
    return None


def resolve_header(template: str, prospect_no) -> str:
    """
    Replace {prospect_no} / {group} placeholders. If none present, append
    the group value so the ID still appears.
    """
    text = template or DEFAULT_HEADER
    replacements = {
        "{prospect_no}": str(prospect_no),
        "{Prospect_No}": str(prospect_no),
        "{group}": str(prospect_no),
        "{GROUP}": str(prospect_no),
    }
    out = text
    for k, v in replacements.items():
        out = out.replace(k, v)
    if not any(k in text for k in replacements):
        out = f"{text} - {prospect_no}"
    return out


def format_dispatch_dates(df: pd.DataFrame) -> pd.DataFrame:
    date_col = find_column(df.columns, ["Dispatch Date", "dispatch_date", "DispatchDate"])
    if date_col:
        df = df.copy()
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%d-%m-%Y")
        df[date_col] = df[date_col].fillna("")
    return df


def drop_empty_columns(df: pd.DataFrame, keep=()) -> tuple:
    """
    Drop columns that are entirely empty (NaN / blank / whitespace).
    Always retains columns listed in `keep` when they exist.
    Returns (cleaned_df, list_of_dropped_column_names).
    """
    keep_set = {str(c) for c in keep}
    dropped = []
    keep_cols = []
    for col in df.columns:
        if col in keep_set:
            keep_cols.append(col)
            continue
        series = df[col]
        # Treat NaN and blank strings as empty
        nonempty = series.dropna().astype(str).str.strip()
        nonempty = nonempty[nonempty != ""]
        nonempty = nonempty[~nonempty.str.lower().isin(["nan", "none", "nat"])]
        if len(nonempty) == 0:
            dropped.append(col)
        else:
            keep_cols.append(col)
    return df[keep_cols].copy(), dropped


def write_group_excel(group_df, excel_path, header_text, wide_col_name):
    """Write one styled Excel file for a prospect group."""
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        group_df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.book.active

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        header = ws.cell(row=1, column=1, value=header_text)
        header.font = Font(bold=True, size=20)
        header.alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )
                cell.border = border

        # Header row (row 2) names → find wide column index
        wide_col_idx = None
        if wide_col_name:
            for cell in ws[2]:
                if cell.value is not None and str(cell.value).strip().lower() == wide_col_name.strip().lower():
                    wide_col_idx = cell.column
                    break
            if wide_col_idx is None:
                # partial match
                for cell in ws[2]:
                    if cell.value and wide_col_name.strip().lower() in str(cell.value).strip().lower():
                        wide_col_idx = cell.column
                        break

        for i in range(1, ws.max_column + 1):
            col_letter = get_column_letter(i)
            values = [
                len(str(cell.value))
                for cell in ws[col_letter][1:]
                if cell.value is not None
            ]
            max_length = max(values, default=10)
            if wide_col_idx is not None and i == wide_col_idx:
                ws.column_dimensions[col_letter].width = WIDE_WIDTH
            else:
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        ws.row_dimensions[1].height = 35
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 45


def excel_to_pdf(excel_app, excel_path, pdf_path):
    wb = excel_app.Workbooks.Open(os.path.abspath(excel_path))
    try:
        ws = wb.ActiveSheet
        ws.PageSetup.Orientation = 2  # xlLandscape
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    finally:
        wb.Close(False)


def run_pipeline(
    input_file,
    out_dir,
    group_col,
    header_template,
    wide_col,
    batch_size,
    log_fn,
    progress_fn,
):
    excel_dir = os.path.join(out_dir, "output_excels")
    pdf_dir = os.path.join(out_dir, "output_pdfs")
    batch_dir = os.path.join(out_dir, "output_batches")
    for d in (excel_dir, pdf_dir, batch_dir):
        os.makedirs(d, exist_ok=True)

    log_fn(f"Reading → {input_file}")
    data = pd.read_excel(input_file, engine="openpyxl")
    data.columns = [str(c).strip() for c in data.columns]

    if group_col not in data.columns:
        # try fuzzy
        found = find_column(data.columns, [group_col])
        if not found:
            raise ValueError(
                f"Group column '{group_col}' not found. "
                f"Available: {', '.join(data.columns)}"
            )
        group_col = found
        log_fn(f"Using group column → {group_col}")

    before_cols = list(data.columns)
    data, dropped_global = drop_empty_columns(data, keep=(group_col,))
    if dropped_global:
        log_fn(f"Dropped empty columns ({len(dropped_global)}): {', '.join(dropped_global)}")
    else:
        log_fn(f"Columns → {len(before_cols)} (none empty)")

    # Resolve wide column against remaining headers
    wide_resolved = None
    if wide_col:
        wide_resolved = find_column(
            data.columns,
            [wide_col, "Address", "address", "Ref No", "ref no", "ref_no"],
        )
        if wide_resolved:
            log_fn(f"Wide column (width {WIDE_WIDTH}) → {wide_resolved}")
        else:
            log_fn(f"⚠ Wide column '{wide_col}' not found — using auto widths only")

    data = format_dispatch_dates(data)
    groups = list(data.groupby(group_col, sort=False))
    total = len(groups)
    if total == 0:
        raise ValueError("No groups found in the Excel file.")

    log_fn(f"Groups   → {total} unique '{group_col}' values")
    log_fn(f"Header   → {header_template}\n")

    try:
        import win32com.client as win32
    except ImportError as e:
        raise RuntimeError(
            "pywin32 is required for Excel→PDF export. pip install pywin32"
        ) from e

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    created = 0
    errors = []
    try:
        for i, (prospect_no, group) in enumerate(groups, 1):
            if pd.isna(prospect_no) or str(prospect_no).strip() == "":
                log_fn(f"⏭ Skipping blank {group_col} group")
                continue

            safe_name = re.sub(r'[<>:"/\\|?*]', "_", str(prospect_no).strip())
            excel_path = os.path.join(excel_dir, f"{safe_name}.xlsx")
            pdf_path = os.path.join(pdf_dir, f"{safe_name}.pdf")

            g = group.copy()
            # Drop columns empty for this prospect only (keep group col)
            g, dropped_group = drop_empty_columns(g, keep=(group_col,))
            wide_for_group = wide_resolved if wide_resolved in g.columns else None
            g.insert(0, "SrNo", range(1, len(g) + 1))
            header_text = resolve_header(header_template, prospect_no)

            try:
                write_group_excel(g, excel_path, header_text, wide_for_group)
                excel_to_pdf(excel, excel_path, pdf_path)
                created += 1
                extra = f", dropped {len(dropped_group)} empty cols" if dropped_group else ""
                log_fn(f"✅ [{i}/{total}] {safe_name} ({len(g)} rows{extra})")
            except Exception as e:
                errors.append((safe_name, str(e)))
                log_fn(f"❌ [{i}/{total}] {safe_name} — {e}")

            progress_fn(i / total * 0.7)
    finally:
        try:
            excel.Quit()
        except Exception:
            pass

    # Ordered PDFs by first appearance of group value
    prospect_order = data[group_col].dropna().unique()
    ordered_pdfs = [
        os.path.join(pdf_dir, f"{re.sub(r'[<>:\"/\\\\|?*]', '_', str(p).strip())}.pdf")
        for p in prospect_order
    ]
    ordered_pdfs = [p for p in ordered_pdfs if os.path.exists(p)]

    if not ordered_pdfs:
        raise ValueError("No PDFs were created.")

    log_fn(f"\nMerging {len(ordered_pdfs)} PDFs in batches of {batch_size}…")
    batch_files = []
    for i in range(0, len(ordered_pdfs), batch_size):
        batch = ordered_pdfs[i:i + batch_size]
        batch_output = os.path.join(batch_dir, f"batch_{i // batch_size + 1}.pdf")
        merger = PdfMerger()
        for pdf in batch:
            merger.append(pdf)
        merger.write(batch_output)
        merger.close()
        batch_files.append(batch_output)
        log_fn(f"  Batch → {os.path.basename(batch_output)} ({len(batch)} files)")
        progress_fn(0.7 + 0.2 * ((i // batch_size + 1) / max(1, (len(ordered_pdfs) + batch_size - 1) // batch_size)))

    final_output = os.path.join(out_dir, "UCP_format.pdf")
    merger = PdfMerger()
    for bf in batch_files:
        merger.append(bf)
    merger.write(final_output)
    merger.close()
    progress_fn(1.0)

    log_fn(f"\n🏁 Final PDF → {final_output}")
    return {
        "created": created,
        "errors": len(errors),
        "final_pdf": final_output,
        "groups": total,
    }


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("UCP Format PDF")
        self.geometry("760x720")
        self.configure(fg_color=C["bg"])
        self._path = None
        self._columns = []
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color=TINT["bg"], corner_radius=0)
        hdr.pack(fill="x")
        inn = ctk.CTkFrame(hdr, fg_color="transparent")
        inn.pack(padx=28, pady=14)
        ctk.CTkLabel(
            inn, text="📄  UCP Format PDF",
            font=ctk.CTkFont("Segoe UI", 18, "bold"), text_color=C["text"],
        ).pack(anchor="w")
        ctk.CTkLabel(
            inn,
            text="Group Excel by Prospect No · styled sheets · PDF merge",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["muted"],
        ).pack(anchor="w")

        body = ctk.CTkScrollableFrame(
            self, fg_color="transparent", scrollbar_button_color=C["border"],
        )
        body.pack(fill="both", expand=True, padx=24, pady=16)

        banner = ctk.CTkFrame(
            body, fg_color=TINT["bg"], corner_radius=10,
            border_width=1, border_color=C["accent"],
        )
        banner.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(
            banner,
            text="📁  Output → Desktop\\OUTPUT\\UCP_Format_PDF\\<timestamp>\\",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["accent"],
        ).pack(anchor="w", padx=14, pady=8)

        self._sec(body, "Excel file")
        fr = ctk.CTkFrame(body, fg_color="transparent")
        fr.pack(fill="x", pady=(0, 10))
        self._file_lbl = ctk.CTkLabel(
            fr, text="No file selected", font=ctk.CTkFont("Segoe UI", 12),
            text_color=C["muted"], anchor="w",
        )
        self._file_lbl.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(
            fr, text="Browse…", width=90, height=34,
            fg_color=C["card"], hover_color=C["hover"],
            border_color=C["border"], border_width=1,
            text_color=C["text"], command=self._pick,
        ).pack(side="right")

        self._sec(body, "Settings")
        settings = ctk.CTkFrame(
            body, fg_color=C["card"], corner_radius=12,
            border_width=1, border_color=C["border"],
        )
        settings.pack(fill="x", pady=(0, 10))
        settings.columnconfigure(1, weight=1)

        self._group_cb = self._dropdown(
            settings, 0, "Group by column",
            [DEFAULT_GROUP_COL],
            DEFAULT_GROUP_COL,
            "Select the ID column after loading an Excel file (default Prospect_No).",
        )
        self._header_e = self._field(
            settings, 1, "Header text", DEFAULT_HEADER,
            "Use {prospect_no} where the group value should appear.",
        )
        self._wide_cb = self._dropdown(
            settings, 2, "Wide column",
            [DEFAULT_WIDE_COL],
            DEFAULT_WIDE_COL,
            "Column that gets width 70 (default Address). Others auto-size.",
        )
        self._batch_e = self._field(
            settings, 3, "PDF batch size", str(DEFAULT_BATCH),
            "How many PDFs to merge per intermediate batch.",
            last=True,
        )

        self._prog = ctk.CTkProgressBar(
            body, height=8, fg_color=C["card"], progress_color=C["accent"],
        )
        self._prog.pack(fill="x", pady=(4, 6))
        self._prog.set(0)
        self._stat = ctk.CTkLabel(
            body, text="Ready.", font=ctk.CTkFont("Segoe UI", 11),
            text_color=C["muted"], anchor="w",
        )
        self._stat.pack(fill="x", pady=(0, 6))
        self._log = ctk.CTkTextbox(
            body, height=200, font=ctk.CTkFont("Courier New", 11),
            fg_color=C["card"], border_color=C["border"],
            border_width=1, text_color=C["muted"], state="disabled",
        )
        self._log.pack(fill="x", pady=(0, 12))

        self._run_btn = ctk.CTkButton(
            body, text="▶  Generate UCP PDFs",
            font=ctk.CTkFont("Segoe UI", 14, "bold"),
            fg_color=TINT["mid"], hover_color=TINT["bdr"],
            text_color=C["accent"], border_color=C["accent"], border_width=1,
            corner_radius=24, height=46, command=self._start,
        )
        self._run_btn.pack(fill="x", pady=(0, 16))

    def _dropdown(self, parent, row, label, values, default, hint):
        ctk.CTkLabel(
            parent, text=label, font=ctk.CTkFont("Segoe UI", 11),
            text_color=C["muted"], width=130,
        ).grid(row=row * 2, column=0, padx=16, pady=(12 if row == 0 else 4, 0), sticky="w")
        cb = ctk.CTkComboBox(
            parent, values=values, height=34,
            font=ctk.CTkFont("Segoe UI", 12),
            fg_color=C["hover"], border_color=C["border"],
            button_color=TINT["mid"], button_hover_color=TINT["bdr"],
            dropdown_fg_color=C["card"], dropdown_hover_color=C["hover"],
            dropdown_text_color=C["text"], text_color=C["text"],
            state="readonly",
        )
        cb.set(default)
        cb.grid(row=row * 2, column=1, padx=16, pady=(12 if row == 0 else 4, 0), sticky="ew")
        ctk.CTkLabel(
            parent, text=hint, font=ctk.CTkFont("Segoe UI", 10),
            text_color=C["faint"],
        ).grid(
            row=row * 2 + 1, column=0, columnspan=2,
            padx=16, pady=(0, 2), sticky="w",
        )
        return cb

    def _field(self, parent, row, label, default, hint, last=False):
        ctk.CTkLabel(
            parent, text=label, font=ctk.CTkFont("Segoe UI", 11),
            text_color=C["muted"], width=130,
        ).grid(row=row * 2, column=0, padx=16, pady=(12 if row == 0 else 4, 0), sticky="w")
        entry = ctk.CTkEntry(
            parent, height=34, fg_color=C["hover"],
            border_color=C["border"], text_color=C["text"],
        )
        entry.grid(row=row * 2, column=1, padx=16, pady=(12 if row == 0 else 4, 0), sticky="ew")
        entry.insert(0, default)
        ctk.CTkLabel(
            parent, text=hint, font=ctk.CTkFont("Segoe UI", 10),
            text_color=C["faint"],
        ).grid(
            row=row * 2 + 1, column=0, columnspan=2,
            padx=16, pady=(0, 10 if last else 2), sticky="w",
        )
        return entry

    def _sec(self, p, t):
        ctk.CTkLabel(
            p, text=t, font=ctk.CTkFont("Segoe UI", 12, "bold"),
            text_color=C["text"], anchor="w",
        ).pack(fill="x", pady=(10, 2))

    def _write(self, msg):
        self._log.configure(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.configure(state="disabled")

    def _pick(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        self._path = path
        try:
            cols = list(pd.read_excel(path, nrows=0).columns)
            self._columns = [str(c).strip() for c in cols]
            if not self._columns:
                raise ValueError("No columns found in the Excel file.")

            self._group_cb.configure(values=self._columns)
            self._wide_cb.configure(values=self._columns)

            g = find_column(
                self._columns,
                [DEFAULT_GROUP_COL, "prospect_no", "Prospect No", "ProspectNo"],
            )
            self._group_cb.set(g or self._columns[0])

            w = find_column(
                self._columns,
                [DEFAULT_WIDE_COL, "address", "Ref No", "ref_no", "ref no", "Reference"],
            )
            self._wide_cb.set(w or self._columns[0])

            self._file_lbl.configure(
                text=f"{os.path.basename(path)}  ({len(self._columns)} cols)",
                text_color=C["accent"],
            )
        except Exception as e:
            messagebox.showerror("Error", f"Could not read Excel:\n{e}")

    def _start(self):
        if not self._path:
            messagebox.showwarning("Missing", "Select an Excel file first.")
            return
        if not self._columns:
            messagebox.showwarning("Missing", "Load an Excel file so columns appear in the dropdowns.")
            return

        group_col = self._group_cb.get().strip()
        wide_col = self._wide_cb.get().strip()
        header = self._header_e.get().strip() or DEFAULT_HEADER
        if not group_col:
            messagebox.showwarning("Missing", "Select a group-by column.")
            return
        if not wide_col:
            messagebox.showwarning("Missing", "Select a wide column.")
            return
        try:
            batch = int(self._batch_e.get().strip())
            if batch < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid", "Batch size must be a positive number.")
            return

        self._run_btn.configure(state="disabled", text="Processing…")
        self._prog.set(0)
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        threading.Thread(
            target=self._run,
            args=(group_col, header, wide_col, batch),
            daemon=True,
        ).start()

    def _run(self, group_col, header, wide_col, batch):
        out_dir = get_output_dir()

        def log(m):
            self.after(0, lambda x=m: self._write(x))

        def prog(p):
            self.after(0, lambda v=p: self._prog.set(v))

        try:
            stats = run_pipeline(
                self._path, out_dir, group_col, header, wide_col, batch, log, prog,
            )
            self.after(0, lambda: self._stat.configure(
                text=f"Done — {stats['created']} PDFs, {stats['errors']} errors.",
                text_color=C["green"],
            ))
            self.after(0, lambda: subprocess.Popen(["explorer", out_dir]))
            self.after(0, lambda: messagebox.showinfo(
                "Complete",
                f"Created: {stats['created']}\n"
                f"Errors: {stats['errors']}\n\n"
                f"{stats['final_pdf']}",
            ))
        except Exception as e:
            log(f"\nError: {e}")
            self.after(0, lambda: self._stat.configure(text=str(e), text_color=C["red"]))
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.after(0, lambda: self._run_btn.configure(
                state="normal", text="▶  Generate UCP PDFs",
            ))


if __name__ == "__main__":
    App().mainloop()
