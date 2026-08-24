"""
Tool: File Downloader (from Excel)
Download any file type from URL / Excel hyperlink columns.
Optional rename using an ID column. Always uses the first sheet.
"""

import os
import re
import mimetypes
import threading
import subprocess
from datetime import datetime
from urllib.parse import urlparse, unquote, parse_qs

import pandas as pd
import requests
import customtkinter as ctk
from tkinter import filedialog, messagebox

try:
    import openpyxl
except ImportError:
    openpyxl = None

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

DESKTOP  = os.path.join(os.path.expanduser("~"), "Desktop")
BASE_OUT = os.path.join(DESKTOP, "OUTPUT", "File_Downloads")

C = {
    "bg":     "#0a0a0f", "card":   "#16161f", "hover":  "#1e1e2e",
    "border": "#2a2a3d", "text":   "#e8e8f0", "muted":  "#8888aa",
    "faint":  "#44445a", "accent": "#00f5ff", "green":  "#30d158",
    "red":    "#ff375f",
}
TINT = {"bg": "#062d30", "mid": "#0a4a4e", "bdr": "#0d6b70"}

EXT_FROM_MIME = {
    "application/pdf": ".pdf",
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/gif": ".gif",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
    "image/tiff": ".tiff",
    "application/zip": ".zip",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.ms-excel": ".xls",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "text/plain": ".txt",
    "text/csv": ".csv",
}


def get_output_dir():
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    path = os.path.join(BASE_OUT, ts)
    os.makedirs(path, exist_ok=True)
    return path


def unique_path(folder, filename):
    dest = os.path.join(folder, filename)
    if not os.path.exists(dest):
        return dest, filename
    base, ext = os.path.splitext(filename)
    n = 1
    while True:
        name = f"{base}_{n}{ext}"
        dest = os.path.join(folder, name)
        if not os.path.exists(dest):
            return dest, name
        n += 1


def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = name.rstrip(". ")
    return name or "file"


def is_http_url(value: str) -> bool:
    v = (value or "").strip().lower()
    return v.startswith("http://") or v.startswith("https://")


def load_urls_from_excel(excel_path: str, url_col: str, id_col: str | None):
    """
    Read first sheet. Prefer Excel hyperlink target on the URL column when
    present (e.g. cell text 'VIEW POD' with a real link underneath).
    Falls back to cell text if it looks like a URL.
    Returns list of dicts: {row, url, id}.
    """
    rows = []

    if openpyxl is not None and excel_path.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb.worksheets[0]
        headers = {}
        for cell in ws[1]:
            if cell.value is not None:
                headers[str(cell.value).strip()] = cell.column

        if url_col not in headers:
            raise ValueError(
                f"URL column '{url_col}' not found on first sheet. "
                f"Available: {', '.join(headers)}"
            )
        url_ci = headers[url_col]
        id_ci = headers.get(id_col) if id_col else None

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=url_ci)
            url = None
            if cell.hyperlink and cell.hyperlink.target:
                url = str(cell.hyperlink.target).strip()
            else:
                raw = "" if cell.value is None else str(cell.value).strip()
                if is_http_url(raw):
                    url = raw

            rid = None
            if id_ci:
                id_val = ws.cell(row=r, column=id_ci).value
                if id_val is not None and str(id_val).strip():
                    rid = str(id_val).strip()

            if url or rid:
                rows.append({"row": r, "url": url or "", "id": rid})
        return rows

    # Fallback (e.g. .xls) — cell text only, first sheet
    df = pd.read_excel(excel_path, sheet_name=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    if url_col not in df.columns:
        raise ValueError(
            f"URL column '{url_col}' not found. Available: {', '.join(df.columns)}"
        )
    if id_col and id_col not in df.columns:
        raise ValueError(
            f"ID column '{id_col}' not found. Available: {', '.join(df.columns)}"
        )

    for i, row in df.iterrows():
        raw = "" if pd.isna(row.get(url_col)) else str(row.get(url_col)).strip()
        url = raw if is_http_url(raw) else ""
        rid = None
        if id_col:
            v = row.get(id_col)
            if not pd.isna(v) and str(v).strip():
                rid = str(v).strip()
        rows.append({"row": int(i) + 2, "url": url, "id": rid})
    return rows


def guess_ext_from_url(url: str) -> str:
    """Try path, then nested ?file= query param (common for POD gateways)."""
    try:
        parsed = urlparse(url)
        path = unquote(parsed.path)
        ext = os.path.splitext(path)[1].lower()
        if ext and len(ext) <= 6:
            return ext

        qs = parse_qs(parsed.query)
        for key in ("file", "url", "path", "src"):
            if key in qs and qs[key]:
                nested = unquote(qs[key][0])
                nested_path = urlparse(nested).path if "://" in nested else nested
                ext = os.path.splitext(nested_path)[1].lower()
                if ext and len(ext) <= 6:
                    return ext
    except Exception:
        pass
    return ""


def guess_ext_from_response(resp, url: str) -> str:
    # Content-Disposition filename
    cd = resp.headers.get("Content-Disposition") or resp.headers.get("content-disposition") or ""
    m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, re.I)
    if m:
        ext = os.path.splitext(unquote(m.group(1).strip()))[1].lower()
        if ext:
            return ext

    ctype = (resp.headers.get("Content-Type") or "").split(";")[0].strip().lower()
    if ctype in EXT_FROM_MIME:
        return EXT_FROM_MIME[ctype]
    guessed = mimetypes.guess_extension(ctype) if ctype else None
    if guessed:
        return guessed

    return guess_ext_from_url(url) or ".bin"


def filename_from_url(url: str, fallback: str) -> str:
    try:
        parsed = urlparse(url)
        name = os.path.basename(unquote(parsed.path))
        if name and "." in name:
            return sanitize_filename(name)
        qs = parse_qs(parsed.query)
        for key in ("file", "url", "path", "src"):
            if key in qs and qs[key]:
                nested = unquote(qs[key][0])
                nested_path = urlparse(nested).path if "://" in nested else nested
                name = os.path.basename(nested_path)
                if name and "." in name:
                    return sanitize_filename(name)
    except Exception:
        pass
    return sanitize_filename(fallback)


def download_file(url: str, dest_folder: str, rename_id: str | None, log_fn) -> str:
    resp = requests.get(url, timeout=60, stream=True, allow_redirects=True)
    resp.raise_for_status()

    ext = guess_ext_from_response(resp, url)
    if rename_id:
        base = sanitize_filename(rename_id)
        if base.lower().endswith(ext.lower()):
            filename = base
        else:
            filename = f"{base}{ext}"
    else:
        filename = filename_from_url(url, f"file{ext}")
        # Ensure extension present
        if not os.path.splitext(filename)[1]:
            filename = f"{filename}{ext}"

    path, final_name = unique_path(dest_folder, filename)
    with open(path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=64 * 1024):
            if chunk:
                f.write(chunk)
    log_fn(f"  Saved → {final_name}")
    return final_name


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("File Downloader")
        self.geometry("760x700")
        self.configure(fg_color=C["bg"])
        self._path = None
        self._build()

    def _build(self):
        hdr = ctk.CTkFrame(self, fg_color=TINT["bg"], corner_radius=0)
        hdr.pack(fill="x")
        inn = ctk.CTkFrame(hdr, fg_color="transparent")
        inn.pack(padx=28, pady=14)
        ctk.CTkLabel(
            inn, text="📥  File Downloader",
            font=ctk.CTkFont("Segoe UI", 18, "bold"), text_color=C["text"],
        ).pack(anchor="w")
        ctk.CTkLabel(
            inn,
            text="Download any file from Excel URLs / hyperlinks · optional rename",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["muted"],
        ).pack(anchor="w")

        body = ctk.CTkScrollableFrame(
            self, fg_color="transparent", scrollbar_button_color=C["border"],
        )
        body.pack(fill="both", expand=True, padx=24, pady=16)

        banner = ctk.CTkFrame(
            body, fg_color=TINT["bg"], corner_radius=10,
            border_width=1, border_color=C["accent"],
        )
        banner.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(
            banner,
            text="📁  Output → Desktop\\OUTPUT\\File_Downloads\\<timestamp>\\",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["accent"],
        ).pack(anchor="w", padx=14, pady=8)

        self._sec(body, "Excel file (first sheet is used)")
        fr = ctk.CTkFrame(body, fg_color="transparent")
        fr.pack(fill="x", pady=(0, 10))
        self._file_lbl = ctk.CTkLabel(
            fr, text="No file selected", font=ctk.CTkFont("Segoe UI", 12),
            text_color=C["muted"], anchor="w",
        )
        self._file_lbl.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(
            fr, text="Browse…", width=90, height=34,
            fg_color=C["card"], hover_color=C["hover"],
            border_color=C["border"], border_width=1,
            text_color=C["text"], command=self._pick,
        ).pack(side="right")

        self._sec(body, "Columns")
        settings = ctk.CTkFrame(
            body, fg_color=C["card"], corner_radius=12,
            border_width=1, border_color=C["border"],
        )
        settings.pack(fill="x", pady=(0, 10))
        settings.columnconfigure(1, weight=1)

        ctk.CTkLabel(
            settings, text="URL / link column",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["muted"], width=140,
        ).grid(row=0, column=0, padx=16, pady=(14, 8), sticky="w")
        self._url_e = ctk.CTkEntry(
            settings, height=34, fg_color=C["hover"],
            border_color=C["border"], text_color=C["text"],
        )
        self._url_e.grid(row=0, column=1, padx=16, pady=(14, 8), sticky="ew")
        self._url_e.insert(0, "POD IMAGE")

        ctk.CTkLabel(
            settings, text="ID column (rename)",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["muted"], width=140,
        ).grid(row=1, column=0, padx=16, pady=(0, 14), sticky="w")
        self._id_e = ctk.CTkEntry(
            settings, height=34, fg_color=C["hover"],
            border_color=C["border"], text_color=C["text"],
            placeholder_text="e.g. AWB NUMBER",
        )
        self._id_e.grid(row=1, column=1, padx=16, pady=(0, 14), sticky="ew")
        self._id_e.insert(0, "AWB NUMBER")

        opts = ctk.CTkFrame(
            body, fg_color=C["card"], corner_radius=12,
            border_width=1, border_color=C["border"],
        )
        opts.pack(fill="x", pady=(0, 12))
        self._rename_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            opts, text="Rename downloaded files using ID column",
            variable=self._rename_var,
            font=ctk.CTkFont("Segoe UI", 12),
            text_color=C["text"], fg_color=C["accent"],
            hover_color=TINT["bdr"], border_color=C["border"],
            command=self._toggle_rename,
        ).pack(anchor="w", padx=16, pady=(12, 4))
        ctk.CTkLabel(
            opts,
            text="Supports cell hyperlinks (e.g. “VIEW POD”) and plain URL text. "
                 "Downloads PDF, JPG, PNG, and other file types.",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["muted"],
            wraplength=680, justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 12))

        self._prog = ctk.CTkProgressBar(
            body, height=8, fg_color=C["card"], progress_color=C["accent"],
        )
        self._prog.pack(fill="x", pady=(4, 6))
        self._prog.set(0)
        self._stat = ctk.CTkLabel(
            body, text="Ready.", font=ctk.CTkFont("Segoe UI", 11),
            text_color=C["muted"], anchor="w",
        )
        self._stat.pack(fill="x", pady=(0, 6))
        self._log = ctk.CTkTextbox(
            body, height=220, font=ctk.CTkFont("Courier New", 11),
            fg_color=C["card"], border_color=C["border"],
            border_width=1, text_color=C["muted"], state="disabled",
        )
        self._log.pack(fill="x", pady=(0, 12))

        self._run_btn = ctk.CTkButton(
            body, text="▶  Start Download",
            font=ctk.CTkFont("Segoe UI", 14, "bold"),
            fg_color=TINT["mid"], hover_color=TINT["bdr"],
            text_color=C["accent"], border_color=C["accent"], border_width=1,
            corner_radius=24, height=46, command=self._start,
        )
        self._run_btn.pack(fill="x", pady=(0, 16))

    def _sec(self, p, t):
        ctk.CTkLabel(
            p, text=t, font=ctk.CTkFont("Segoe UI", 12, "bold"),
            text_color=C["text"], anchor="w",
        ).pack(fill="x", pady=(10, 2))

    def _toggle_rename(self):
        state = "normal" if self._rename_var.get() else "disabled"
        self._id_e.configure(state=state)

    def _write(self, msg):
        self._log.configure(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.configure(state="disabled")

    def _pick(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")],
        )
        if path:
            self._path = path
            self._file_lbl.configure(
                text=os.path.basename(path), text_color=C["accent"],
            )

    def _start(self):
        if not self._path:
            messagebox.showwarning("Missing", "Select an Excel file first.")
            return
        url_col = self._url_e.get().strip()
        if not url_col:
            messagebox.showwarning("Missing", "Enter the URL / link column name.")
            return
        rename = self._rename_var.get()
        id_col = self._id_e.get().strip() if rename else ""
        if rename and not id_col:
            messagebox.showwarning(
                "Missing",
                "Enter an ID column for renaming, or turn off rename.",
            )
            return

        self._run_btn.configure(state="disabled", text="Downloading…")
        self._prog.set(0)
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")
        threading.Thread(
            target=self._run,
            args=(url_col, id_col if rename else None, rename),
            daemon=True,
        ).start()

    def _run(self, url_col, id_col, rename):
        out_dir = get_output_dir()

        def log(m):
            self.after(0, lambda x=m: self._write(x))

        try:
            log(f"Excel   → {self._path}")
            log(f"Sheet   → first sheet")
            log(f"URL col → {url_col}")
            log(f"Rename  → {'ON (' + id_col + ')' if rename else 'OFF'}")
            log(f"Output  → {out_dir}\n")

            items = load_urls_from_excel(self._path, url_col, id_col)
            total = len(items)
            if total == 0:
                raise ValueError("No rows found on the first sheet.")

            ok = fail = skip = 0
            for i, item in enumerate(items, 1):
                url = item["url"]
                rid = item["id"]
                row_n = item["row"]
                self.after(0, lambda v=i / total: self._prog.set(v))
                self.after(0, lambda a=i, b=total: self._stat.configure(
                    text=f"Processing {a} / {b}…", text_color=C["muted"],
                ))

                if not url or not is_http_url(url):
                    log(f"⚠️  Row {row_n}: no valid URL/hyperlink — skipped")
                    skip += 1
                    continue

                label = rid or f"row_{row_n}"
                log(f"↓ Row {row_n}: {label}")
                try:
                    download_file(
                        url, out_dir,
                        rename_id=rid if rename else None,
                        log_fn=log,
                    )
                    ok += 1
                except Exception as e:
                    log(f"  ❌ {e}")
                    fail += 1

            log("\n========== SUMMARY ==========")
            log(f"Downloaded : {ok}")
            log(f"Failed     : {fail}")
            log(f"Skipped    : {skip}")
            log(f"Folder     : {out_dir}")

            self.after(0, lambda: self._prog.set(1))
            self.after(0, lambda: self._stat.configure(
                text=f"Done — {ok} downloaded, {fail} failed, {skip} skipped.",
                text_color=C["green"],
            ))
            self.after(0, lambda: subprocess.Popen(["explorer", out_dir]))
            self.after(0, lambda: messagebox.showinfo(
                "Complete",
                f"Downloaded: {ok}\nFailed: {fail}\nSkipped: {skip}\n\n{out_dir}",
            ))
        except Exception as e:
            log(f"\nError: {e}")
            self.after(0, lambda: self._stat.configure(text=str(e), text_color=C["red"]))
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.after(0, lambda: self._run_btn.configure(
                state="normal", text="▶  Start Download",
            ))


if __name__ == "__main__":
    App().mainloop()
